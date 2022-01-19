import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter.font import BOLD
from tkinter.ttk import *
from tkinter.constants import CENTER, TOP, DISABLED
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import pyexcel as p
import pyexcel_xls
import pyexcel_xlsx
from datetime import datetime
import pandas as pd
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
import time
from selenium.webdriver.common.by import By
import glob
import winshell
from win32com.client import Dispatch
import pickle
import datetime
from collections import namedtuple
from google_auth_oauthlib.flow import Flow, InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.auth.transport.requests import Request



# delete inconvenient, possible waste files
local_path = os.getcwd()
listardir = os.listdir(local_path)
for item in listardir:
    if item.endswith(".xls"):
        os.remove(os.path.join(local_path, item))
    if item.endswith(".xlsx"):
        os.remove(os.path.join(local_path,item))


# create desktop shortcut
desktop = winshell.desktop()
path = os.path.join(desktop, "Agendador ePROC.lnk")
target = os.path.join(local_path, "Agendador ePROC.exe")
wDir = local_path
icon = os.path.join(local_path, "Agendador ePROC.exe")

if not os.path.exists(path):
    shell = Dispatch('WScript.Shell')
    shortcut = shell.CreateShortCut(path)
    shortcut.Targetpath = target
    shortcut.WorkingDirectory = wDir
    shortcut.IconLocation = icon
    shortcut.save()


# create api service function
def create_service(client_secret_file, api_name, api_version, *scopes, prefix=''):
	CLIENT_SECRET_FILE = client_secret_file
	API_SERVICE_NAME = api_name
	API_VERSION = api_version
	SCOPES = [scope for scope in scopes[0]]
	
	cred = None
	working_dir = os.getcwd()
	token_dir = 'token files'
	pickle_file = f'token_{API_SERVICE_NAME}_{API_VERSION}{prefix}.pickle'

	# check if token dir exists first, if not, create the folder
	if not os.path.exists(os.path.join(working_dir, token_dir)):
		os.mkdir(os.path.join(working_dir, token_dir))

	if os.path.exists(os.path.join(working_dir, token_dir, pickle_file)):
		with open(os.path.join(working_dir, token_dir, pickle_file), 'rb') as token:
			cred = pickle.load(token)

	if not cred or not cred.valid:
		if cred and cred.expired and cred.refresh_token:
			cred.refresh(Request())
		else:
			flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
			cred = flow.run_local_server()

		with open(os.path.join(working_dir, token_dir, pickle_file), 'wb') as token:
			pickle.dump(cred, token)

	try:
		service = build(API_SERVICE_NAME, API_VERSION, credentials=cred)
		print(API_SERVICE_NAME, API_VERSION, 'service created successfully')
		return service
	except Exception as e:
		print(e)
		print(f'Failed to create service instance for {API_SERVICE_NAME}')
		os.remove(os.path.join(working_dir, token_dir, pickle_file))
		return None


# convert datetime function
def convert_to_RFC_datetime(year=1900, month=1, day=1, hour=0, minute=0):
	dt = datetime.datetime(year, month, day, hour, minute, 0).isoformat() + 'Z'
	return dt



CLIENT_SECRET_FILE = "client_secret.json"
API_NAME = 'calendar'
API_VERSION = 'v3'
SCOPES = ['https://www.googleapis.com/auth/calendar']



service = create_service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)

# list calendars
calendar_list = service.calendarList().list(pageToken=None,maxResults=10).execute()


# delete already existing calendar
for calendar_list_entry in calendar_list['items']:  
    if 'Intimações ePROC' in calendar_list_entry['summary']:
        id = calendar_list_entry['id'] 
        service.calendars().delete(calendarId=id).execute()


# create new calendar
calendar_body = {
    'summary': 'Intimações ePROC',
    'timeZone': 'America/Sao_Paulo',
}
service.calendars().insert(body=calendar_body).execute()


# list calendars
calendar_list = service.calendarList().list(pageToken=None,maxResults=6).execute()

# get calendar id
for calendar_list_entry in calendar_list['items']:
    if 'Intimações ePROC' in calendar_list_entry['summary']:
        id = calendar_list_entry['id'] 



# center window
def center_window(width=860,height=640):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    root.geometry('%dx%d+%d+%d' % (width, height, x, y))


# close program confirmation
def on_closeroot():
    close = messagebox.askokcancel("Confirmação", "Tem certeza que deseja fechar o programa?")
    if close:
        root.destroy()

# main function
def runcode():
    # style
    status = tk.Label(text="Carregando... Aguarde",wraplength=200,font=('',9,'bold'),bg='white')
    status.pack()
    status.place(relx=0.5, rely=0.7, anchor=CENTER)
    status.update_idletasks()

    button1.configure(state = 'disabled')
    button1.update_idletasks()
    time.sleep(1)
    button1.configure(state = 'enabled')
    status.destroy()


    # ------------------------------------- SANTA CATARINA ----------------------------------

    # web browser download xls file
    # configuring web browser
    options = Options()
    login = "INSERT_LOGIN_HERE"                                     # INSERIR LOGIN DO EPROC
    passwd = "INSERT_PASSWORD_HERE"                                 # INSERIR SENHA DO EPROC
    options.add_argument("start-maximized")
    options.add_argument("--headless")
    preferences = {"download.default_directory": local_path,
                    "directory_upgrade": True}
    options.add_experimental_option("prefs", preferences)                   

    # open web browser
    driverpath = "/chromedriver.exe"
    driver = webdriver.Chrome(executable_path = local_path + driverpath, options=options)
    driver.get("https://eproc1g.tjsc.jus.br/eproc/externo_controlador.php?acao=principal")
    time.sleep(1)

    # login
    loginxpath = driver.find_element(By.XPATH, '//*[@id="txtUsuario"]')
    loginxpath.send_keys(login)

    passwxpath = driver.find_element(By.XPATH, '//*[@id="pwdSenha"]')
    passwxpath.send_keys(passwd)

    loginbuttonxpath = driver.find_element(By.XPATH, '//*[@id="sbmEntrar"]')
    loginbuttonxpath.click()

    # get to file webpage
    scbuttonxpath = driver.find_element(By.XPATH, '//*[@id="tr0"]')
    scbuttonxpath.click()
    time.sleep(2)

    intimacoesxpath = driver.find_element(By.XPATH, '//*[@id="conteudoCitacoesIntimacoesSC"]/div[2]/table/tbody/tr[1]/td[2]/a')
    intimacoesxpath.click()
    time.sleep(1)

    newURl = driver.window_handles[1]
    driver.switch_to.window(newURl)

    # download file
    gerarplanilhaid = driver.find_element(By.ID, 'sbmPlanilha')
    gerarplanilhaid.click()
    time.sleep(2)
    driver.quit()

    # rename downloaded file
    filename = glob.glob('*.xls')
    filename = ''.join(filename)
    finalfilename = filename[:16] + '.xls'
    os.rename(filename, finalfilename)

    # convert to xlsx
    p.save_book_as(file_name= finalfilename,
                dest_file_name='intimacaosc.xlsx')


    # load file, sheet
    wb = load_workbook('intimacaosc.xlsx')
    ws = wb.active


    # formatting excel chart
    for i in range(0,2):
        ws.delete_rows(1)

    ws.insert_rows(1)

    for i in range(0,2):
        ws.delete_cols(1)

    for i in range(0,5):
        ws.delete_cols(2)


    # row number variable
    max_rows = ws.max_row-1

    # append list variables
    subj = []
    date = []
    filetime = []


    # copy subject
    for i in range(1,max_rows+2):
        subj.append(ws.cell(row = i, column = 1).value)

    # paste into description
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 7).value = subj[i-1]


    # copy date
    for i in range(1,max_rows+2):
        date.append(ws.cell(row = i, column = 2).value)

    # stringify date
    i = 0
    while i < max_rows:
        i += 1
        date[i] = date[i].strftime("%x")

    # paste date 
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 2).value = date[i-1]

    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 4).value = date[i-1]



    # fill start time
    a = 0
    for i in range(1,max_rows+2):
        if i < 18:
            ws.cell(row = i, column = 3).value = "{}:00".format(i)
        elif i >= 18:
            a += 1
            ws.cell(row = i, column = 3).value = "{}:00".format(a)
            if a >= 17:
                a = 0

    # copy start time
    for i in range(1,max_rows+2):
        filetime.append(ws.cell(row = i, column = 3).value)

    # paste end time
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 5).value = filetime[i-1]



    # fill all day event
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 6).value = "FALSE"  


    # fill location
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 8).value = "ePROC-SC" 


    # fill private
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 9).value = "TRUE"    



    # header
    cabecalho = ['Subject', 'Start Date', 'Start Time', 'End Date', 'End Time', 'All Day Event', 'Description', 'Location', 'Private']
    for col in range(0,9):
        char = chr(65 + col)
        ws[char + '1'] = cabecalho[col]



    # save file
    wb.save('sc.xlsx')


    # insert events to google calendar function
    def insert_events(location, color):
        for i in range(0,max_rows):
            adjust_timezone = 3
            adjust_timezone_endtime = 4
            event_request_body = {
                'start':{
                    'dateTime': convert_to_RFC_datetime(int(ano[i]), int(mes[i]), int(dia[i]), int(tempo[i]) + adjust_timezone, 0),
                    'timeZone': 'America/Sao_Paulo',
                },
                'end':{
                    'dateTime': convert_to_RFC_datetime(int(ano[i]), int(mes[i]), int(dia[i]), int(tempo[i]) + adjust_timezone_endtime, 0),
                    'timeZone': 'America/Sao_Paulo',
                },
                'summary': subj[i],
                'description': subj[i],
                'location': location,
                'colorId': color,
                #'attendees':[
                #    {
                #        'email': '',
                #        'optional': False,
                #        'responseStatus': 'accepted',
                #    }
                #],
                #'reminders': {
                #    'useDefault': False,
                #    'overrides':[
                #        {'method': 'email', 'minutes': 30},
                #    ]
                #}
            }
            service.events().insert(calendarId=id, body=event_request_body).execute()




    # load xlsx file containing the events
    wb = load_workbook('sc.xlsx')
    ws = wb.active

    # row number variable
    max_rows = ws.max_row-1

    # append lists
    subj = []
    date = []
    filetime = []
    tempo = []
    mes = []
    dia = []
    ano = []


    # copy subject
    for i in range(2,max_rows+2):
        subj.append(ws.cell(row = i, column = 1).value)


    # copy date
    for i in range(1,max_rows+2):
        date.append(ws.cell(row = i, column = 2).value)


    # copy start time
    for i in range(1,max_rows+2):
        filetime.append(ws.cell(row = i, column = 3).value)



    # format time
    for i in range(1,max_rows+1):
        tempo.append(filetime[i])
    tempo = [x[:-3] for x in tempo]


    # format month
    for i in range(1,max_rows+1):
        mes.append(date[i])
    mes = [x[:-6] for x in mes]


    # format day
    for i in range(1,max_rows+1):
        dia.append(date[i])
    dia = [x[3:-3] for x in dia]


    # format year
    for i in range(1,max_rows+1):
        ano.append(date[i])
    ano = ['20' + x[6:] for x in ano]


    # insert events
    sc = "ePROC-SC"
    insert_events(sc, 7)    


    # delete no longer needed files
    os.remove(finalfilename)
    os.remove('intimacaosc.xlsx')


    # -------------------------------------- PARANA ----------------------------------------------------

    # web browser download xls file
    # open web browser
    driver = webdriver.Chrome(executable_path = local_path + driverpath, options=options)
    driver.get("https://eproc1g.tjsc.jus.br/eproc/externo_controlador.php?acao=principal")
    time.sleep(1)

    # login
    loginxpath = driver.find_element(By.XPATH, '//*[@id="txtUsuario"]')
    loginxpath.send_keys(login)

    passwxpath = driver.find_element(By.XPATH, '//*[@id="pwdSenha"]')
    passwxpath.send_keys(passwd)

    loginbuttonxpath = driver.find_element(By.XPATH, '//*[@id="sbmEntrar"]')
    loginbuttonxpath.click()

    # get to file webpage
    prbuttonxpath = driver.find_element(By.XPATH, '//*[@id="tr1"]')
    prbuttonxpath.click()
    time.sleep(2)

    intimacoesxpath = driver.find_element(By.XPATH, '//*[@id="conteudoCitacoesIntimacoesSC"]/div[2]/table/tbody/tr[1]/td[2]/a')
    intimacoesxpath.click()
    time.sleep(1)

    newURl = driver.window_handles[1]
    driver.switch_to.window(newURl)

    # download file
    gerarplanilhaid = driver.find_element(By.ID, 'sbmPlanilha')
    gerarplanilhaid.click()
    time.sleep(2)
    driver.quit()

    # rename downloaded file
    filename = glob.glob('*.xls')
    filename = ''.join(filename)
    finalfilename = filename[:16] + '.xls'
    os.rename(filename, finalfilename)


    # convert to xlsx
    p.save_book_as(file_name= finalfilename,
                dest_file_name='intimacaopr.xlsx')


    # load file, sheet
    wb = load_workbook('intimacaopr.xlsx')
    ws = wb.active


    # formatting excel chart
    for i in range(0,2):
        ws.delete_rows(1)

    ws.insert_rows(1)

    for i in range(0,2):
        ws.delete_cols(1)

    for i in range(0,5):
        ws.delete_cols(2)


    # row number variable
    max_rows = ws.max_row-1

    # append list variables
    subj = []
    date = []
    filetime = []


    # copy subject
    for i in range(1,max_rows+2):
        subj.append(ws.cell(row = i, column = 1).value)

    # paste into description
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 7).value = subj[i-1]



    # copy date
    for i in range(1,max_rows+2):
        date.append(ws.cell(row = i, column = 2).value)

    # stringify date
    i = 0
    while i < max_rows:
        i += 1
        date[i] = date[i].strftime("%x")

    # paste date 
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 2).value = date[i-1]

    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 4).value = date[i-1]



    # fill start time
    a = 0
    for i in range(1,max_rows+2):
        if i < 18:
            ws.cell(row = i, column = 3).value = "{}:00".format(i)
        elif i >= 18:
            a += 1
            ws.cell(row = i, column = 3).value = "{}:00".format(a)
            if a >= 17:
                a = 0

    # copy start time
    for i in range(1,max_rows+2):
        filetime.append(ws.cell(row = i, column = 3).value)

    # paste end time
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 5).value = filetime[i-1]



    # fill all day event
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 6).value = "FALSE"  


    # fill location
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 8).value = "ePROC-PR" 


    # fill private
    for i in range(1,max_rows+2):
        ws.cell(row = i, column = 9).value = "TRUE"    



    # header
    cabecalho = ['Subject', 'Start Date', 'Start Time', 'End Date', 'End Time', 'All Day Event', 'Description', 'Location', 'Private']
    for col in range(0,9):
        char = chr(65 + col)
        ws[char + '1'] = cabecalho[col]



    # save file
    wb.save('pr.xlsx')




    # load xlsx file containing the events
    wb = load_workbook('pr.xlsx')
    ws = wb.active

    # row number variable
    max_rows = ws.max_row-1

    # append lists
    subj = []
    date = []
    filetime = []
    tempo = []
    mes = []
    dia = []
    ano = []

    # copy subject
    for i in range(2,max_rows+2):
        subj.append(ws.cell(row = i, column = 1).value)


    # copy date
    for i in range(1,max_rows+2):
        date.append(ws.cell(row = i, column = 2).value)


    # copy start time
    for i in range(1,max_rows+2):
        filetime.append(ws.cell(row = i, column = 3).value)



    # format time
    for i in range(1,max_rows+1):
        tempo.append(filetime[i])
    tempo = [x[:-3] for x in tempo]


    # format month
    for i in range(1,max_rows+1):
        mes.append(date[i])
    mes = [x[:-6] for x in mes]


    # format day
    for i in range(1,max_rows+1):
        dia.append(date[i])
    dia = [x[3:-3] for x in dia]


    # format year
    for i in range(1,max_rows+1):
        ano.append(date[i])
    ano = ['20' + x[6:] for x in ano]

    # insert events
    pr = "ePROC-PR"
    insert_events(pr, 11)


    # delete no longer needed files
    os.remove(finalfilename)
    os.remove('intimacaopr.xlsx')

    # ----------------------------------------------------------------------------------------------

    # dynamic GUI
    status1 = tk.Label(text="Concluído!",wraplength=200,font=('',9,'bold'),bg='white')
    status1.pack()
    status1.place(relx=0.5, rely=0.7, anchor=CENTER)


    # confirmation
    def on_closetop():
        status = tk.Label(text="As intimações foram importadas e se encontram no Google Agenda!",wraplength=200,font=('',9,'bold'),bg='white')
        status.pack()
        status.place(relx=0.5, rely=0.71, anchor=CENTER)
        status.update_idletasks()
        close = messagebox.showinfo("Informação", "As intimações foram adicionadas no Google Agenda com sucesso!")
        if close:
            close2 = messagebox.askyesno("Sair","Deseja sair do programa?")
            if close2:
                root.destroy()
            else:
                status.destroy()
                status.update_idletasks()

    # end program
    def endprogram():
        status1.destroy()
        on_closetop() 


    closeconfirmation = messagebox.showinfo("Sucesso!", "O programa foi executado com sucesso!")
    if closeconfirmation:
        endprogram()
    

    #
    #
    #
    # python -m PyInstaller --onedir --windowed --icon=icone.ico --name="Agendador ePROC" Agendador.py
    #
    #
    #

# static GUI
root = tk.Tk()
center_window(860, 640)
root.title("Agendador ePROC")
bg = PhotoImage(file = "background.png")
background_label = Label(root, image=bg)
background_label.place(relx=0.5,rely=0.5,anchor=CENTER)
root.protocol("WM_DELETE_WINDOW", on_closeroot)
iconFile = 'icone.ico'
root.iconbitmap(default=iconFile)

# menu instructions
info = tk.Label(text="Clique no botão abaixo para iniciar o programa e aguarde alguns instantes!",wraplength=200,font=('',13),bg='white')
info.pack()
info.place(relx=0.5, rely=0.4, anchor=CENTER)

# welcoming message
status = tk.Label(text="Seja bem-vindo!",wraplength=200,font=('',15,'bold'),bg='white')
status.pack()
status.place(relx=0.5, rely=0.52, anchor=CENTER)

# credits
feitopor = tk.Label(text="Programa criado por: Gianluca Notari Magnabosco da Silva",font=('',7),bg="white")
feitopor.pack()
feitopor.place(relx=0.85, rely=0.98, anchor=CENTER)

# run code button
st = Style()
st.configure('W.TButton', background='white', foreground='black', font=('Open Sans',11))
button1 = Button(root, style='W.TButton', text='Clique aqui para iniciar o programa',command=runcode,width=27.75)
button1.pack()
button1.place(relx=0.5, rely=0.65, anchor=CENTER)

# end of GUI
root.mainloop()